import PySimpleGUI as sg
import openpyxl
import json
from collections import defaultdict
from datetime import datetime

# Defina os valores padrão para mês e ano
mes_padrao = 2
ano_padrao = 2024

# Função para ler dados da base de dados dos funcionários
def ler_dados_funcionarios(arquivo):
    workbook = openpyxl.load_workbook(arquivo)
    sheet = workbook.active

    funcionarios = defaultdict(lambda: defaultdict(int))
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data, nome_completo, empresa, torre, andar = row
        if empresa == "Conbras_Itau":
            funcionarios[data.day]["ItauBBA"] += 1  # Adicionar à categoria ItauBBA
        elif empresa in ["Controller", "Hagana", "HeatingCooling", "Innova",
                         "Não Atribuido", "NetPark", "Temon", "ThyssenKrupp", "Verzanni_WTNU"]:
            funcionarios[data.day]["Terceiros WTNU"] += 1  # Adicionar à categoria Terceiros WTNU
        else:
            funcionarios[data.day][empresa] += 1

    return funcionarios

# Função para ler dados da base de dados dos visitantes

def ler_dados_visitantes(arquivo):
    workbook = openpyxl.load_workbook(arquivo)
    sheet = workbook.active

    visitantes = defaultdict(lambda: defaultdict(int))
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_str, _, nome, visitado, _, _, _ = row
        data_str = str(data_str)  # Garantir que data_str seja uma string
        
        # Verificar se data_str possui um formato válido
        if len(data_str) < 10:
            continue  # Ignorar esta linha e passar para a próxima
        
        try:
            data = datetime.strptime(data_str, "%Y-%m-%d %H:%M:%S")  # Converter a string de data para um objeto de data
        except ValueError:
            continue  # Ignorar esta linha e passar para a próxima

        if not visitado or visitado == "Selecionar...":
            visitantes[data.day]["Visitante WTNU"] += 1
        else:
            visitado = visitado if visitado != "Conbras_Itau" else "ItauBBA"
            if visitado in ["Controller", "Hagana", "HeatingCooling", "Innova",
                          "Temon", "ThyssenKrupp", "Verzanni_WTNU"]:
                visitantes[data.day]["Visitante WTNU"] += 1
            else:
                visitantes[data.day][visitado] += 1

    return visitantes

def escrever_dados_relatorio(workbook, dados_funcionarios, dados_visitantes, arquivo_saida, mes, ano):
    sheet = workbook.active
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escrever cabeçalhos (incluir as colunas adicionais)
    header_row = ['Dia'] + ordem_empresas + ['Visitante WTNU', 'Total', 'Data', 'Pessoas', '%']
    sheet.append(header_row)


    # Escrever dados
    for dia in sorted(dados_funcionarios.keys()):
        funcionarios_dia = dados_funcionarios[dia]
        visitantes_dia = dados_visitantes[dia]

        # Calcular o total de pessoas no dia (funcionários + visitantes)
        total_pessoas = sum(funcionarios_dia.values()) + sum(visitantes_dia.values())

        # Calcular a porcentagem de ocupação com uma casa decimal
        porcentagem_ocupacao = (total_pessoas / 8931) * 100

        row_data = [f'{dia:02d}/{mes:02d}/{ano}']  # Formato dia/mes/ano
        for empresa in ordem_empresas:
            if empresa.startswith("Visitante "):
                visitante_empresa = empresa[len("Visitante "):]
                row_data.append(visitantes_dia[visitante_empresa])
            else:
                row_data.append(funcionarios_dia[empresa])

        # Incluir o valor da coluna "Visitante WTNU", "Total" e as novas colunas
        row_data.append(visitantes_dia["Visitante WTNU"])
        row_data.append(total_pessoas)
        row_data.append(f'{dia:02d}/{mes:02d}/{ano}')  # Formato dia/mes/ano
        row_data.append(total_pessoas)
        row_data.append(f'{porcentagem_ocupacao:.1f}%')  # Formatando a porcentagem com uma casa decimal e o símbolo '%'
        sheet.append(row_data)

    # Calcular a soma de funcionários em cada empresa
    soma_funcionarios_empresas = defaultdict(int)
    for funcionarios_dia in dados_funcionarios.values():
        for empresa, quantidade in funcionarios_dia.items():
            soma_funcionarios_empresas[empresa] += quantidade

    # Calcular a soma de visitantes em cada empresa
    soma_visitantes_empresas = defaultdict(int)
    for visitantes_dia in dados_visitantes.values():
        for empresa, quantidade in visitantes_dia.items():
            soma_visitantes_empresas[empresa] += quantidade

    # Adicionar linha "ENTRADA" com a soma de funcionários e visitantes em cada empresa
    entrada_row = ["ENTRADA"]
    for empresa in ordem_empresas:
        if empresa.startswith("Visitante "):
            visitante_empresa = empresa[len("Visitante "):]
            entrada_row.append(soma_visitantes_empresas[visitante_empresa])  # Soma de visitantes
        else:
            entrada_row.append(soma_funcionarios_empresas[empresa])  # Soma de funcionários
    entrada_row.append(soma_visitantes_empresas["Visitante WTNU"])  # Soma de visitantes WTNU
    entrada_row.append("")  # Coluna "Total"
    entrada_row.append("")  # Coluna de data (dia/mes/ano)
    entrada_row.append("")  # Coluna de total de pessoas
    entrada_row.append("")  # Coluna de porcentagem de ocupação
    sheet.append(entrada_row)

    # Salvar o arquivo de saída
    workbook.save(arquivo_saida)

# Função para adicionar uma nova empresa à lista
def adicionar_empresa(lista_empresas, nova_empresa):
    lista_empresas.append(nova_empresa)

# Função para remover uma empresa da lista
def remover_empresa(lista_empresas, empresa_removida):
    lista_empresas.remove(empresa_removida)

# Função para salvar a lista de empresas no arquivo de configuração
def salvar_configuracao(lista_empresas):
    with open('config.json', 'w') as f:
        json.dump(lista_empresas, f)

def carregar_configuracao():
    try:
        with open('config.json', 'r') as f:
            data = f.read()
            if data:
                return json.loads(data)
            else:
                return []
    except FileNotFoundError:
        return []

def mover_empresa_para_cima(lista_empresas, index):
    if index > 0:
        lista_empresas[index], lista_empresas[index - 1] = lista_empresas[index - 1], lista_empresas[index]

# Função para reordenar empresas para baixo
def mover_empresa_para_baixo(lista_empresas, index):
    if index < len(lista_empresas) - 1:
        lista_empresas[index], lista_empresas[index + 1] = lista_empresas[index + 1], lista_empresas[index]

# Defina a lista ordem_empresas globalmente e carregue as configurações
ordem_empresas = carregar_configuracao()

#########################Definir a interface gráfica#########################

# Defina o layout da segunda tela de configurações
layout_configuracoes = [
    [sg.Text('Configurações de Empresas')],
    [sg.Text('Empresas:')],
    [sg.Listbox(values=ordem_empresas, size=(30, 6), key='empresas')],
    [sg.Button('Adicionar Empresa'), sg.Button('Remover Empresa')],
    [sg.Button('Mover Para Cima'), sg.Button('Mover Para Baixo')],
    [sg.Button('Salvar Configurações'), sg.Button('Fechar')]
]

# Defina o layout da tela principal
layout_principal = [
    [sg.Text('Selecione os arquivos de entrada:')],
    [sg.FileBrowse('Selecionar base de funcionários', key='funcionarios')],
    [sg.FileBrowse('Selecionar base de visitantes', key='visitantes')],
    [sg.Text('Selecione o mês da data:'),
     sg.Combo([str(i) for i in range(1, 13)], default_value=str(mes_padrao), key='mes')],
    [sg.Text('Selecione o ano da data:'),
     sg.InputText(default_text=str(ano_padrao), key='ano')],
    [sg.Button('Gerar Relatório'), sg.Button('Configurações'), sg.Button('Sair')]
]

# Crie a janela principal
window = sg.Window('Gerador de Relatório', layout_principal)

# Loop principal
while True:
    event, values = window.read()

    if event == sg.WINDOW_CLOSED or event == 'Sair':
        break
    elif event == 'Gerar Relatório':
        arquivo_funcionarios = values['funcionarios']
        arquivo_visitantes = values['visitantes']
        mes_escolhido = int(values['mes'])
        ano_escolhido = int(values['ano'])

        if arquivo_funcionarios and arquivo_visitantes:
            dados_funcionarios = ler_dados_funcionarios(arquivo_funcionarios)
            dados_visitantes = ler_dados_visitantes(arquivo_visitantes)
            
            workbook = openpyxl.Workbook()  # Criar um novo workbook
            escrever_dados_relatorio(workbook, dados_funcionarios, dados_visitantes, 'relatorio_acesso.xlsx', mes_escolhido, ano_escolhido)

            sg.popup('Relatório gerado com sucesso!', title='Sucesso')
        else:
            sg.popup('Por favor, selecione os arquivos de entrada.', title='Erro')

        pass
    elif event == 'Configurações':
        # Abra a janela de configurações
        window_config = sg.Window('Configurações', layout_configuracoes)
        while True:
            event_config, values_config = window_config.read()

            if event_config == sg.WINDOW_CLOSED or event_config == 'Fechar':
                break
            elif event_config == 'Adicionar Empresa':
                # Use Popup para obter uma nova empresa do usuário
                nova_empresa = sg.popup_get_text('Digite o nome da nova empresa:')
                
                # Verifique se o usuário não cancelou a entrada
                if nova_empresa:
                    adicionar_empresa(ordem_empresas, nova_empresa)
                    window_config['empresas'].update(ordem_empresas)
                    sg.popup(f'Empresa "{nova_empresa}" adicionada com sucesso!', title='Sucesso')
                else:
                    sg.popup("Operação cancelada pelo usuário.", title='Aviso')
            elif event_config == 'Remover Empresa':
                empresa_removida = values_config['empresas'][0] if values_config['empresas'] else None
                if empresa_removida:
                    remover_empresa(ordem_empresas, empresa_removida)
                    window_config['empresas'].update(ordem_empresas)
                    sg.popup(f'Empresa "{empresa_removida}" removida com sucesso!', title='Sucesso')
            elif event_config == 'Mover Para Cima':
                selected_index = window_config['empresas'].get_indexes()[0]
                if selected_index is not None and selected_index > 0:
                    mover_empresa_para_cima(ordem_empresas, selected_index)
                    window_config['empresas'].update(ordem_empresas)
            elif event_config == 'Mover Para Baixo':
                selected_index = window_config['empresas'].get_indexes()[0]
                if selected_index is not None and selected_index < len(ordem_empresas) - 1:
                    mover_empresa_para_baixo(ordem_empresas, selected_index)
                    window_config['empresas'].update(ordem_empresas)
            elif event_config == 'Salvar Configurações':
                salvar_configuracao(ordem_empresas)
                sg.popup('Configurações salvas com sucesso!', title='Sucesso')

        window_config.close()

# Feche a janela principal
window.close()